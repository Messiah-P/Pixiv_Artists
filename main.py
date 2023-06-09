import requests
import traversal
from get_single_pic import SinglePic
from search import Artist
from multi_thread_download import multi_download
import numpy as np
from openpyxl import load_workbook
from config import artists_info, ARTISTS_DIR, LOGO_PIXIV, HEAD_BARK
from log import log_output

def get_illust(artists_list, existed_pids):
     # 通过输入构造目录
     bark_new = '--更新画作--'
     bark_null = '--未更新--'
     for i in range(len(artists_list)): #遍历行
         artist_name = artists_list[i][0]
         artist_id = artists_list[i][1]
         log_output(f"画师:{artist_name} - {artist_id}")
         pic = SinglePic()
         pic.mkdir(ARTISTS_DIR)
         # 创建对象获取该画师插画列表
         log_output(f"正在获取{artist_name}的插画...")
         artist = Artist(artist_id)
         illust_ids = list(artist.get_illust_ids(existed_pids))
         if len(illust_ids) != 0:
             log_output(f"{artist_name}共有{str(len(illust_ids))}件作品未下载, 开始下载{artist_name}的作品...")
             bark_info = ('画师: ' + artist_name + '\n' +'更新数: ' + str(len(illust_ids)))
             bark_new = bark_new + '\n' + bark_info
             #创建多线程下载
             log_output(f"创建多线程下载...")
             multi_download(illust_ids, 0, ARTISTS_DIR)
         else:
             log_output(f"{artist_name}没有作品更新。")
             bark_null = bark_null + '\n' + artist_name
     bark = bark_new + '\n' + bark_null
     ret = requests.get('%s/Pixiv画师更新/%s?icon=%s&group=画师' % (HEAD_BARK, bark, LOGO_PIXIV))
     print('\n下载结束啦')

if __name__ == '__main__':
    wb = load_workbook(artists_info)  # Pixiv工作表文件
    sh = wb['作者信息']  # 作者信息
    artists_list = np.empty(shape=(0, 2))  # 生成0行2列数组
    # 获取作者信息
    for i in range(sh.max_row -1): # 获取最大行
        user_info_row = []
        for col in sh.iter_cols(min_row=2):
            user_info_row.append(col[i].value)
        artists_list = np.row_stack((artists_list, user_info_row))
    log_output(f"{artists_list}")

    existed_pids = traversal.all_pids(ARTISTS_DIR)
    get_illust(artists_list, existed_pids)
