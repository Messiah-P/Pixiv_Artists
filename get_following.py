import numpy as np

from following.crawler import Crawler
from openpyxl import load_workbook
from config import artists_info
from log import log_output


class GetFollowing(object):
    def __init__(self):
        self.artists_info = artists_info

    def update_userlist(self):
        wb = load_workbook(self.artists_info)  # Pixiv工作表文件
        sh = wb['作者信息']  # 作者信息
        c = Crawler()
        following_list = c.run()

        artists_list = np.empty(shape=(0, 2))  # 生成0行2列数组
        existing_data = np.empty(shape=(0, 2))  # 生成0行2列数组
        # 获取作者信息
        for i in range(sh.max_row - 1):  # 获取最大行
            user_info_row = []
            for col in sh.iter_cols(min_row=2):
                user_info_row.append(col[i].value)
            existing_data = np.row_stack((existing_data, user_info_row))
        log_output(f"{existing_data}")

        # 初始化一个新的NumPy数组
        new_data = []

        # 遍历原始数据
        for item in following_list:
            uid = item['uid']
            userName = item['userName']

            # 检查uid是否已存在于现有数据中
            if not np.isin(uid, existing_data[:, 1]):
                # 如果uid不在现有数据中，添加userName和uid到新数据中
                print("新关注画师：",userName,uid)
                new_data.append([userName, uid])

        # 将新数据添加到现有数据中
        if new_data:
            new_data = np.array(new_data)
            existing_data = np.concatenate((existing_data, new_data), axis=0)
        #print(existing_data)

        # 将现有数据（包括新数据）写回Excel
        for i, row in enumerate(artists_list):
            for j, value in enumerate(row):
                sh.cell(row=i + 2, column=j + 1, value=value)

        # 保存Excel文件
        wb.save(artists_info)

        return existing_data
